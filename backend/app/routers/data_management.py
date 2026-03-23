"""
Excel 模板下载、数据导入、按依赖清空（不导出数据）
"""
from __future__ import annotations

import io
import re
from datetime import date, datetime, time
from typing import Any, Dict, List, Optional, Set, Tuple

from dateutil import parser as date_parser
from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.utils.datetime import from_excel
from pydantic import BaseModel
from sqlalchemy.orm import Session

from ..database import get_db
from .. import models
from .auth import require_admin, require_auth
from .orders import _allocate_order_number, replace_order_operations_from_routing

router = APIRouter()


def _require_location_master(db: Session, code: Any) -> str:
    c = str(code or "").strip()
    if not c:
        raise ValueError("位置代码必填（请先在「位置」主数据维护）")
    if not db.query(models.Location).filter(models.Location.code == c).first():
        raise ValueError(f"位置主数据不存在: {c}")
    return c


DATA_TYPES = [
    "locations",
    "work_centers",
    "resources",
    "shifts",
    "products",
    "routings",
    "routing_operations",
    "setup_groups",
    "product_setup_groups",
    "setup_matrix",
    "production_orders",
]

# 清空执行顺序（先删子表）
CLEAR_STEPS: List[Tuple[str, Any]] = [
    ("operations", models.Operation),
    ("production_orders", models.ProductionOrder),
    ("product_setup_groups", models.ProductSetupGroup),
    ("setup_matrix", models.SetupMatrix),
    ("routing_operations", models.RoutingOperation),
    ("routings", models.Routing),
    ("shifts", models.Shift),
    ("resources", models.Resource),
    ("work_centers", models.WorkCenter),
    ("products", models.Product),
    ("setup_groups", models.SetupGroup),
    ("locations", models.Location),
]

# 勾选某类型时需一并清空的类型（含自身），满足外键与业务依赖
CLEAR_EXPANSION: Dict[str, Set[str]] = {
    "locations": {"locations"},
    "operations": {"operations"},
    "production_orders": {"operations", "production_orders"},
    "product_setup_groups": {"product_setup_groups"},
    "setup_matrix": {"setup_matrix"},
    "routing_operations": {"operations", "routing_operations"},
    "routings": {"operations", "routing_operations", "routings"},
    "shifts": {"shifts"},
    "resources": {
        "operations",
        "production_orders",
        "product_setup_groups",
        "setup_matrix",
        "routing_operations",
        "routings",
        "shifts",
        "resources",
    },
    "work_centers": {
        "operations",
        "production_orders",
        "product_setup_groups",
        "setup_matrix",
        "routing_operations",
        "routings",
        "shifts",
        "resources",
        "work_centers",
    },
    "products": {
        "operations",
        "production_orders",
        "product_setup_groups",
        "routing_operations",
        "routings",
        "products",
    },
    "setup_groups": {"product_setup_groups", "setup_matrix", "setup_groups"},
}


def _norm_locale(locale: Optional[str], accept_language: Optional[str]) -> str:
    raw = (locale or accept_language or "zh-CN").strip()
    low = raw.lower()
    if low.startswith("en"):
        return "en-US"
    return "zh-CN"


def _norm_header(s: Any) -> str:
    if s is None:
        return ""
    t = str(s).strip().lower()
    t = re.sub(r"\s+", "_", t)
    return t


# (internal_key, zh_label, en_label, extra_aliases)
FIELD_SPECS: Dict[str, List[Tuple[str, str, str, Tuple[str, ...]]]] = {
    "work_centers": [
        ("code", "编码", "Code", ("work_center_code",)),
        ("name", "名称", "Name", ()),
        ("description", "描述", "Description", ()),
    ],
    "resources": [
        ("code", "资源编码", "Resource code", ("resource_code",)),
        ("name", "资源名称", "Resource name", ()),
        ("work_center_code", "工作中心编码", "Work center code", ("wc_code",)),
        ("location", "位置", "Location", ()),
        ("operating_start", "开始", "Start", ("start_time", "start")),
        ("operating_end", "结束", "End", ("end_time", "end")),
        ("operating_break", "休息时段", "Break period", ("break_time", "break_period")),
        ("utilization_percent", "利用效率(%)", "Utilization %", ("utilization",)),
        ("production_hours", "生产时间(小时)", "Production hours (h)", ()),
        ("capacity_value", "能力值", "Capacity value", ("capacity",)),
        ("finite_planning", "有限计划", "Finite planning", ()),
        ("is_bottleneck", "瓶颈资源", "Bottleneck", ()),
        ("efficiency", "效率系数", "Efficiency", ()),
        ("capacity_per_day", "每日产能(小时)", "Capacity per day (h)", ("daily_capacity",)),
        ("timezone", "时区", "Timezone", ()),
        ("factory_calendar", "工厂日历", "Factory calendar", ()),
        ("planning_group", "计划组", "Planning group", ()),
        ("description", "描述", "Description", ()),
    ],
    "locations": [
        ("code", "位置代码", "Location code", ()),
        ("description", "位置描述", "Description", ()),
    ],
    "shifts": [
        ("resource_code", "资源编码", "Resource code", ()),
        ("shift_code", "班次编码", "Shift code", ()),
        ("shift_name", "班次名称", "Shift name", ()),
        ("start_time", "开始时间", "Start time", ()),
        ("end_time", "结束时间", "End time", ()),
        ("break_time", "休息(分钟)", "Break (min)", ("break",)),
        ("location", "位置", "Location", ()),
    ],
    "products": [
        ("code", "产品编码", "Product code", ("product_code",)),
        ("name", "产品描述", "Product description", ("product_description",)),
        ("unit", "基本单位", "Base unit", ("base_unit",)),
        ("product_type", "产品类型", "Product type", ()),
        ("location", "地点", "Location", ()),
        ("location_name", "地点名称", "Location name", ()),
        ("mrp_controller", "MRP控制员", "MRP controller", ()),
        ("mrp_controller_name", "MRP控制员名称", "MRP controller name", ()),
        ("deletion_flag", "删除标记", "Deletion flag", ("deleted",)),
        ("description", "描述", "Description", ()),
    ],
    "routings": [
        ("code", "工艺路线编码", "Routing code", ("routing_code", "编码")),
        ("name", "工艺路线名称", "Routing name", ("名称",)),
        ("product_code", "产品编码", "Product code", ()),
        ("version", "版本", "Version", ()),
        ("is_active", "是否启用", "Active", ("active",)),
        ("location", "位置", "Location", ()),
        ("description", "工艺路线描述", "Routing description", ("描述",)),
        ("sequence", "工序序号", "Operation sequence", ("seq", "op_seq", "序号")),
        ("operation_name", "工序名称", "Operation name", ("工序名称", "op_name")),
        ("work_center_code", "工作中心编码", "Work center code", ("wc_code",)),
        ("resource_code", "资源编码", "Resource code", ()),
        ("setup_time", "准备时间(小时)", "Setup time (h)", ()),
        ("run_time_per_unit", "单件工时(小时)", "Run time per unit (h)", ("run_time",)),
        ("operation_description", "工序描述", "Operation description", ()),
    ],
    "routing_operations": [
        ("routing_code", "工艺路线编码", "Routing code", ()),
        ("sequence", "序号", "Sequence", ("seq",)),
        ("name", "名称", "Name", ()),
        ("work_center_code", "工作中心编码", "Work center code", ("wc_code",)),
        ("resource_code", "资源编码", "Resource code", ()),
        ("setup_time", "准备时间(小时)", "Setup time (h)", ()),
        ("run_time_per_unit", "单件工时(小时)", "Run time per unit (h)", ("run_time",)),
        ("description", "描述", "Description", ()),
    ],
    "setup_groups": [
        ("code", "编码", "Code", ("setup_group_code",)),
        ("name", "名称", "Name", ()),
        ("description", "描述", "Description", ()),
    ],
    "product_setup_groups": [
        ("product_code", "产品编码", "Product code", ()),
        ("setup_group_code", "切换组编码", "Setup group code", ()),
        ("work_center_code", "工作中心编码", "Work center code", ("wc_code",)),
    ],
    "setup_matrix": [
        ("from_setup_group_code", "来源切换组编码", "From setup group code", ("from_code",)),
        ("to_setup_group_code", "目标切换组编码", "To setup group code", ("to_code",)),
        ("resource_code", "资源编码", "Resource code", ()),
        ("work_center_code", "工作中心编码", "Work center code", ("wc_code",)),
        ("changeover_time", "切换时间(小时)", "Changeover time (h)", ()),
        ("description", "描述", "Description", ()),
        ("location", "位置", "Location", ()),
    ],
    "production_orders": [
        ("order_number", "订单号", "Order number", ("order_no",)),
        ("product_code", "产品编码", "Product code", ()),
        ("quantity", "数量", "Quantity", ()),
        ("due_date", "交期", "Due date", ()),
        ("earliest_start", "最早开工", "Earliest start", ()),
        ("priority", "优先级", "Priority", ()),
        ("order_type", "订单类型", "Order type", ()),
        ("description", "描述", "Description", ()),
    ],
}


def _alias_to_internal(data_type: str) -> Dict[str, str]:
    m: Dict[str, str] = {}
    for internal, _zh, _en, extras in FIELD_SPECS[data_type]:
        m[_norm_header(internal)] = internal
        m[_norm_header(_zh)] = internal
        m[_norm_header(_en)] = internal
        for ex in extras:
            m[_norm_header(ex)] = internal
    return m


def _header_labels(data_type: str, loc: str) -> List[str]:
    use_en = loc == "en-US"
    return [
        (en if use_en else zh)
        for (_internal, zh, en, _extras) in FIELD_SPECS[data_type]
    ]


def _sample_row(data_type: str) -> List[Any]:
    samples = {
        "work_centers": ["WC01", "装配线1", "示例工作中心"],
        "resources": [
            "R01",
            "机床1",
            "",
            "1020",
            "09:00:00",
            "18:00:00",
            "00:00:00",
            100,
            8.0,
            "",
            "是",
            "否",
            1.0,
            8.0,
            "CET",
            "01",
            "A",
            "",
        ],
        "locations": ["1020", "示例工厂"],
        "shifts": ["R01", "D1", "白班", "08:00", "17:00", 60, "1001"],
        "products": [
            "P01",
            "产品A",
            "PCS",
            "FERT",
            "1020",
            "Frankfurt Plant",
            "001",
            "",
            "否",
            "",
        ],
        "routings": [
            "RT-P01",
            "P01标准工艺",
            "P01",
            "1.0",
            1,
            "1020",
            "",
            10,
            "工序10",
            "WC01",
            "R01",
            0.5,
            0.1,
            "",
        ],
        "routing_operations": ["RT-P01", 10, "工序10", "WC01", "R01", 0.5, 0.1, ""],
        "setup_groups": ["SG01", "黑色件", ""],
        "product_setup_groups": ["P01", "SG01", "WC01"],
        # 工作中心、资源均留空表示厂级全局矩阵，与页面「全局」视图一致
        "setup_matrix": ["SG01", "SG02", "", "", 0.5, "", "1001"],
        "production_orders": ["", "P01", 100, "2025-12-31", "", 5, "planned", ""],
    }
    return samples[data_type]


def _build_workbook(data_type: str, loc: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = data_type[:31]
    ws.append(_header_labels(data_type, loc))
    if data_type == "routings":
        row = _sample_row(data_type)
        ws.append(row)
        ws.append(
            [
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
                row[5],
                row[6],
                20,
                "工序20" if not loc or loc != "en-US" else "Operation 20",
                row[9],
                row[10],
                0,
                0.2,
                "",
            ]
        )
    else:
        ws.append(_sample_row(data_type))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _parse_optional_bool(v: Any, default: bool) -> bool:
    if v is None or (isinstance(v, str) and v.strip() == ""):
        return default
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(int(v))
    s = str(v).strip().lower()
    if s in ("1", "true", "yes", "y", "是", "启用"):
        return True
    if s in ("0", "false", "no", "n", "否", "停用"):
        return False
    return default


def _normalize_unit_cell(v: Any) -> str:
    s = str(v or "").strip()
    if not s:
        return "PCS"
    m = re.search(r"\(([^)]+)\)\s*$", s)
    return (m.group(1).strip() if m else s)[:20]


def _parse_bool_cell(v: Any) -> int:
    if v is None or v == "":
        return 1
    if isinstance(v, bool):
        return 1 if v else 0
    if isinstance(v, (int, float)):
        return 1 if int(v) != 0 else 0
    s = str(v).strip().lower()
    if s in ("1", "true", "yes", "y", "是", "启用", "active"):
        return 1
    if s in ("0", "false", "no", "n", "否", "停用", "inactive"):
        return 0
    return 1


def _parse_float(v: Any, default: Optional[float] = None) -> Optional[float]:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _parse_int(v: Any, default: Optional[int] = None) -> Optional[int]:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _parse_datetime(v: Any) -> Optional[datetime]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date) and not isinstance(v, datetime):
        return datetime.combine(v, time(23, 59, 0))
    if type(v) in (int, float):
        try:
            x = float(v)
            # Excel 日期序列（常见约 30000–60000）；openpyxl 用 from_excel 转换
            if 200 < x < 1_000_000:
                return from_excel(x)
        except (ValueError, TypeError, OSError):
            pass
    s = str(v).strip()
    if not s:
        return None
    try:
        return date_parser.parse(s)
    except (ValueError, TypeError):
        return None


def _map_row(ws, row_idx: int, col_map: Dict[str, int]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for key, col in col_map.items():
        out[key] = ws.cell(row=row_idx, column=col).value
    return out


def _resolve_work_center_from_resource(
    db: Session, resource_id: Optional[int], work_center_id: Optional[int]
) -> Optional[int]:
    """推导工艺路线工序的工作中心：资源上无关联工作中心时允许为 NULL（仅指定资源）。"""
    if resource_id is not None:
        r = db.query(models.Resource).filter(models.Resource.id == resource_id).first()
        if not r:
            raise ValueError("资源不存在")
        if r.work_center_id is not None:
            return r.work_center_id
        if work_center_id is not None:
            wc = db.query(models.WorkCenter).filter(models.WorkCenter.id == work_center_id).first()
            if not wc:
                raise ValueError("工作中心不存在")
            return work_center_id
        return None
    if work_center_id is None:
        raise ValueError("请填写工作中心编码或资源编码")
    wc = db.query(models.WorkCenter).filter(models.WorkCenter.id == work_center_id).first()
    if not wc:
        raise ValueError("工作中心不存在")
    return work_center_id


def _expand_clear_types(requested: Set[str]) -> List[str]:
    out: Set[str] = set()
    for t in requested:
        out |= CLEAR_EXPANSION.get(t, {t})
    order_keys = [k for k, _ in CLEAR_STEPS]
    return [k for k in order_keys if k in out]


class ClearBody(BaseModel):
    data_types: Optional[List[str]] = None
    all: bool = False


def _import_work_centers(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            code = str(row.get("code") or "").strip()
            name = str(row.get("name") or "").strip()
            if not code or not name:
                raise ValueError("编码与名称必填")
            desc = row.get("description")
            desc = None if desc is None or str(desc).strip() == "" else str(desc).strip()
            existing = db.query(models.WorkCenter).filter(models.WorkCenter.code == code).first()
            if existing:
                existing.name = name
                existing.description = desc
            else:
                db.add(models.WorkCenter(code=code, name=name, description=desc))
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _import_resources(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            code = str(row.get("code") or "").strip()
            name = str(row.get("name") or "").strip()
            wcc = str(row.get("work_center_code") or "").strip()
            if not code or not name:
                raise ValueError("资源编码、资源名称必填")
            wc_id: Optional[int] = None
            if wcc:
                wc = db.query(models.WorkCenter).filter(models.WorkCenter.code == wcc).first()
                if wc:
                    wc_id = wc.id

            up = _parse_float(row.get("utilization_percent"))
            eff_in = _parse_float(row.get("efficiency"))
            if up is not None:
                eff = max(0.0, min(2.0, up / 100.0))
            elif eff_in is not None:
                eff = eff_in
            else:
                eff = 1.0

            cpd = _parse_float(row.get("capacity_per_day"))
            cv = _parse_float(row.get("capacity_value"))
            ph = _parse_float(row.get("production_hours"))
            if cpd is not None:
                cap = cpd
            elif cv is not None:
                cap = cv
            elif ph is not None:
                cap = ph
            else:
                cap = 8.0

            def _sopt(k: str) -> Optional[str]:
                v = row.get(k)
                if v is None or str(v).strip() == "":
                    return None
                return str(v).strip()

            desc = _sopt("description")
            existing = db.query(models.Resource).filter(models.Resource.code == code).first()
            payload = dict(
                name=name,
                work_center_id=wc_id,
                capacity_per_day=cap,
                efficiency=eff,
                description=desc,
                location=_sopt("location"),
                operating_start=_sopt("operating_start"),
                operating_end=_sopt("operating_end"),
                operating_break=_sopt("operating_break"),
                utilization_percent=up,
                production_hours=ph,
                capacity_value=cv,
                finite_planning=_parse_optional_bool(row.get("finite_planning"), True),
                is_bottleneck=_parse_optional_bool(row.get("is_bottleneck"), False),
                timezone=_sopt("timezone"),
                factory_calendar=_sopt("factory_calendar"),
                planning_group=_sopt("planning_group"),
            )
            if payload.get("location"):
                _require_location_master(db, payload["location"])
            if existing:
                for k, v in payload.items():
                    setattr(existing, k, v)
            else:
                db.add(models.Resource(code=code, **payload))
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _import_locations(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            code = str(row.get("code") or "").strip()
            if not code:
                raise ValueError("位置代码必填")
            desc = row.get("description")
            desc = None if desc is None or str(desc).strip() == "" else str(desc).strip()
            existing = db.query(models.Location).filter(models.Location.code == code).first()
            if existing:
                existing.description = desc
            else:
                db.add(models.Location(code=code, description=desc))
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _import_shifts(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            rc = str(row.get("resource_code") or "").strip()
            sc = str(row.get("shift_code") or "").strip()
            sn = str(row.get("shift_name") or "").strip()
            st = str(row.get("start_time") or "").strip()
            et = str(row.get("end_time") or "").strip()
            if not rc or not sc or not sn or not st or not et:
                raise ValueError("资源编码、班次编码、名称、开始/结束时间必填")
            res = db.query(models.Resource).filter(models.Resource.code == rc).first()
            if not res:
                raise ValueError(f"资源不存在: {rc}")
            loc = _require_location_master(db, row.get("location"))
            rl = str(res.location or "").strip()
            if rl and loc != rl:
                raise ValueError("班次位置须与资源位置一致")
            br = _parse_int(row.get("break_time"), 0) or 0
            existing = (
                db.query(models.Shift)
                .filter(models.Shift.resource_id == res.id, models.Shift.shift_code == sc)
                .first()
            )
            if existing:
                existing.shift_name = sn
                existing.start_time = st
                existing.end_time = et
                existing.break_time = br
                existing.location = loc
            else:
                db.add(
                    models.Shift(
                        resource_id=res.id,
                        shift_code=sc,
                        shift_name=sn,
                        start_time=st,
                        end_time=et,
                        break_time=br,
                        location=loc,
                    )
                )
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _import_products(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            code = str(row.get("code") or "").strip()
            name = str(row.get("name") or "").strip()
            if not code or not name:
                raise ValueError("产品编码与产品描述必填")

            def _sopt(k: str) -> Optional[str]:
                v = row.get(k)
                if v is None or str(v).strip() == "":
                    return None
                return str(v).strip()

            unit = _normalize_unit_cell(row.get("unit"))
            desc = _sopt("description")
            existing = db.query(models.Product).filter(models.Product.code == code).first()
            payload = dict(
                name=name,
                description=desc,
                unit=unit,
                product_type=_sopt("product_type"),
                location=_sopt("location"),
                location_name=_sopt("location_name"),
                mrp_controller=_sopt("mrp_controller"),
                mrp_controller_name=_sopt("mrp_controller_name"),
                deletion_flag=_parse_optional_bool(row.get("deletion_flag"), False),
            )
            if payload.get("location"):
                _require_location_master(db, payload["location"])
            if existing:
                for k, v in payload.items():
                    setattr(existing, k, v)
            else:
                db.add(models.Product(code=code, **payload))
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _import_routings(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            code = str(row.get("code") or "").strip()
            name = str(row.get("name") or "").strip()
            pc = str(row.get("product_code") or "").strip()
            if not code or not name or not pc:
                raise ValueError("工艺路线编码、工艺路线名称、产品编码必填")
            prod = db.query(models.Product).filter(models.Product.code == pc).first()
            if not prod:
                raise ValueError(f"产品不存在: {pc}")
            ver = str(row.get("version") or "1.0").strip() or "1.0"
            ia = _parse_bool_cell(row.get("is_active"))
            rloc = _require_location_master(db, row.get("location"))
            desc = row.get("description")
            desc = None if desc is None or str(desc).strip() == "" else str(desc).strip()
            existing = db.query(models.Routing).filter(models.Routing.code == code).first()
            if existing:
                existing.name = name
                existing.product_id = prod.id
                existing.version = ver
                existing.is_active = ia
                existing.description = desc
                existing.location = rloc
                routing = existing
            else:
                routing = models.Routing(
                    code=code,
                    name=name,
                    product_id=prod.id,
                    version=ver,
                    is_active=ia,
                    description=desc,
                    location=rloc,
                )
                db.add(routing)
                db.flush()

            op_seq = _parse_int(row.get("sequence"))
            if op_seq is None:
                ok += 1
                continue

            op_name = str(row.get("operation_name") or "").strip()
            if not op_name:
                raise ValueError("工序名称必填")
            wcc = str(row.get("work_center_code") or "").strip()
            res_code = str(row.get("resource_code") or "").strip()
            rid: Optional[int] = None
            wcid: Optional[int] = None
            if res_code:
                res = db.query(models.Resource).filter(models.Resource.code == res_code).first()
                if not res:
                    raise ValueError(f"资源不存在: {res_code}")
                rid = res.id
                wcid = res.work_center_id
            elif wcc:
                wc = db.query(models.WorkCenter).filter(models.WorkCenter.code == wcc).first()
                if not wc:
                    raise ValueError(f"工作中心不存在: {wcc}")
                wcid = wc.id
            else:
                raise ValueError("请填写工作中心编码或资源编码")
            setup = _parse_float(row.get("setup_time"), 0.0) or 0.0
            run_u = _parse_float(row.get("run_time_per_unit"))
            if run_u is None:
                raise ValueError("单件工时必填")
            op_desc = row.get("operation_description")
            op_desc = None if op_desc is None or str(op_desc).strip() == "" else str(op_desc).strip()
            wcid = _resolve_work_center_from_resource(db, rid, wcid)
            op_existing = (
                db.query(models.RoutingOperation)
                .filter(
                    models.RoutingOperation.routing_id == routing.id,
                    models.RoutingOperation.sequence == op_seq,
                )
                .first()
            )
            if op_existing:
                op_existing.name = op_name
                op_existing.work_center_id = wcid
                op_existing.resource_id = rid
                op_existing.setup_time = setup
                op_existing.run_time_per_unit = run_u
                op_existing.description = op_desc
            else:
                db.add(
                    models.RoutingOperation(
                        routing_id=routing.id,
                        sequence=op_seq,
                        name=op_name,
                        work_center_id=wcid,
                        resource_id=rid,
                        setup_time=setup,
                        run_time_per_unit=run_u,
                        description=op_desc,
                    )
                )
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _import_routing_operations(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            rc = str(row.get("routing_code") or "").strip()
            seq = _parse_int(row.get("sequence"))
            nm = str(row.get("name") or "").strip()
            if not rc or seq is None or not nm:
                raise ValueError("工艺路线编码、序号、名称必填")
            routing = db.query(models.Routing).filter(models.Routing.code == rc).first()
            if not routing:
                raise ValueError(f"工艺路线不存在: {rc}")
            wcc = str(row.get("work_center_code") or "").strip()
            res_code = str(row.get("resource_code") or "").strip()
            rid: Optional[int] = None
            wcid: Optional[int] = None
            if res_code:
                res = db.query(models.Resource).filter(models.Resource.code == res_code).first()
                if not res:
                    raise ValueError(f"资源不存在: {res_code}")
                rid = res.id
                wcid = res.work_center_id
            elif wcc:
                wc = db.query(models.WorkCenter).filter(models.WorkCenter.code == wcc).first()
                if not wc:
                    raise ValueError(f"工作中心不存在: {wcc}")
                wcid = wc.id
            else:
                raise ValueError("请填写工作中心编码或资源编码")
            setup = _parse_float(row.get("setup_time"), 0.0) or 0.0
            run_u = _parse_float(row.get("run_time_per_unit"))
            if run_u is None:
                raise ValueError("单件工时必填")
            desc = row.get("description")
            desc = None if desc is None or str(desc).strip() == "" else str(desc).strip()
            wcid = _resolve_work_center_from_resource(db, rid, wcid)
            existing = (
                db.query(models.RoutingOperation)
                .filter(
                    models.RoutingOperation.routing_id == routing.id,
                    models.RoutingOperation.sequence == seq,
                )
                .first()
            )
            if existing:
                existing.name = nm
                existing.work_center_id = wcid
                existing.resource_id = rid
                existing.setup_time = setup
                existing.run_time_per_unit = run_u
                existing.description = desc
            else:
                db.add(
                    models.RoutingOperation(
                        routing_id=routing.id,
                        sequence=seq,
                        name=nm,
                        work_center_id=wcid,
                        resource_id=rid,
                        setup_time=setup,
                        run_time_per_unit=run_u,
                        description=desc,
                    )
                )
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _import_setup_groups(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            code = str(row.get("code") or "").strip()
            name = str(row.get("name") or "").strip()
            if not code or not name:
                raise ValueError("编码与名称必填")
            desc = row.get("description")
            desc = None if desc is None or str(desc).strip() == "" else str(desc).strip()
            existing = db.query(models.SetupGroup).filter(models.SetupGroup.code == code).first()
            if existing:
                existing.name = name
                existing.description = desc
            else:
                db.add(models.SetupGroup(code=code, name=name, description=desc))
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _import_product_setup_groups(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            pc = str(row.get("product_code") or "").strip()
            sgc = str(row.get("setup_group_code") or "").strip()
            if not pc or not sgc:
                raise ValueError("产品编码与切换组编码必填")
            prod = db.query(models.Product).filter(models.Product.code == pc).first()
            if not prod:
                raise ValueError(f"产品不存在: {pc}")
            sg = db.query(models.SetupGroup).filter(models.SetupGroup.code == sgc).first()
            if not sg:
                raise ValueError(f"切换组不存在: {sgc}")
            wcc = str(row.get("work_center_code") or "").strip()
            wcid: Optional[int] = None
            if wcc:
                wc = db.query(models.WorkCenter).filter(models.WorkCenter.code == wcc).first()
                if wc:
                    wcid = wc.id
                # 未匹配到工作中心时不报错，按未指定工作中心导入（与切换矩阵导入一致）
            q = db.query(models.ProductSetupGroup).filter(
                models.ProductSetupGroup.product_id == prod.id,
            )
            if wcid is None:
                q = q.filter(models.ProductSetupGroup.work_center_id.is_(None))
            else:
                q = q.filter(models.ProductSetupGroup.work_center_id == wcid)
            existing = q.first()
            if existing:
                existing.setup_group_id = sg.id
            else:
                db.add(
                    models.ProductSetupGroup(
                        product_id=prod.id,
                        setup_group_id=sg.id,
                        work_center_id=wcid,
                    )
                )
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _import_setup_matrix(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            fc = str(row.get("from_setup_group_code") or "").strip()
            tc = str(row.get("to_setup_group_code") or "").strip()
            if not fc or not tc:
                raise ValueError("来源/目标切换组编码必填")
            fg = db.query(models.SetupGroup).filter(models.SetupGroup.code == fc).first()
            tg = db.query(models.SetupGroup).filter(models.SetupGroup.code == tc).first()
            if not fg or not tg:
                raise ValueError("切换组不存在")
            res_code = str(row.get("resource_code") or "").strip()
            wcc = str(row.get("work_center_code") or "").strip()
            rid: Optional[int] = None
            wcid: Optional[int] = None
            if res_code:
                res = db.query(models.Resource).filter(models.Resource.code == res_code).first()
                if not res:
                    raise ValueError(f"资源不存在: {res_code}")
                rid = res.id
            if wcc:
                wc = db.query(models.WorkCenter).filter(models.WorkCenter.code == wcc).first()
                if wc:
                    wcid = wc.id
                # 未匹配到工作中心时不报错，按未指定工作中心导入（与资源等主数据解耦）
            ct = _parse_float(row.get("changeover_time"))
            if ct is None:
                raise ValueError("切换时间必填")
            desc = row.get("description")
            desc = None if desc is None or str(desc).strip() == "" else str(desc).strip()
            mloc = _require_location_master(db, row.get("location"))
            q = db.query(models.SetupMatrix).filter(
                models.SetupMatrix.from_setup_group_id == fg.id,
                models.SetupMatrix.to_setup_group_id == tg.id,
                models.SetupMatrix.location == mloc,
            )
            if rid is None:
                q = q.filter(models.SetupMatrix.resource_id.is_(None))
            else:
                q = q.filter(models.SetupMatrix.resource_id == rid)
            if wcid is None:
                q = q.filter(models.SetupMatrix.work_center_id.is_(None))
            else:
                q = q.filter(models.SetupMatrix.work_center_id == wcid)
            existing = q.first()
            if existing:
                existing.changeover_time = ct
                existing.description = desc
                existing.location = mloc
            else:
                db.add(
                    models.SetupMatrix(
                        from_setup_group_id=fg.id,
                        to_setup_group_id=tg.id,
                        resource_id=rid,
                        work_center_id=wcid,
                        changeover_time=ct,
                        description=desc,
                        location=mloc,
                    )
                )
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _parse_order_type(v: Any) -> str:
    if v is None or str(v).strip() == "":
        return models.OrderType.PLANNED.value
    s = str(v).strip().lower()
    if s in ("production", "prod", "生产", "生产订单"):
        return models.OrderType.PRODUCTION.value
    return models.OrderType.PLANNED.value


def _import_production_orders(db: Session, rows: List[Tuple[int, Dict[str, Any]]]) -> Tuple[int, int, List[dict]]:
    ok, fail = 0, 0
    errors: List[dict] = []
    for row_idx, row in rows:
        try:
            on = str(row.get("order_number") or "").strip()
            pc = str(row.get("product_code") or "").strip()
            qty = _parse_float(row.get("quantity"))
            due = _parse_datetime(row.get("due_date"))
            es = _parse_datetime(row.get("earliest_start"))
            pr = _parse_int(row.get("priority"), 5) or 5
            ot = _parse_order_type(row.get("order_type"))
            miss: List[str] = []
            if not pc:
                miss.append("产品编码")
            if qty is None:
                miss.append("数量")
            if due is None:
                miss.append("交期")
            if miss:
                raise ValueError("、".join(miss) + "必填（交期可为 Excel 日期单元格或文本）")
            if not on:
                db.flush()
                on = _allocate_order_number(db, ot)
            prod = db.query(models.Product).filter(models.Product.code == pc).first()
            if not prod:
                raise ValueError(f"产品不存在: {pc}")
            desc = row.get("description")
            desc = None if desc is None or str(desc).strip() == "" else str(desc).strip()
            existing = (
                db.query(models.ProductionOrder)
                .filter(models.ProductionOrder.order_number == on)
                .first()
            )
            if existing:
                existing.product_id = prod.id
                existing.quantity = qty
                existing.due_date = due
                existing.earliest_start = es
                existing.priority = pr
                existing.order_type = ot
                existing.description = desc
                db_order = existing
            else:
                db_order = models.ProductionOrder(
                    order_number=on,
                    product_id=prod.id,
                    quantity=qty,
                    due_date=due,
                    earliest_start=es,
                    priority=pr,
                    order_type=ot,
                    status=(
                        models.OrderStatus.SCHEDULED.value
                        if ot == models.OrderType.PRODUCTION.value
                        else models.OrderStatus.CREATED.value
                    ),
                    description=desc,
                )
                db.add(db_order)
            db.flush()
            db_order.location = _require_location_master(db, prod.location)
            replace_order_operations_from_routing(db, db_order)
            ok += 1
        except Exception as e:
            fail += 1
            errors.append({"row": row_idx, "message": str(e)})
    return ok, fail, errors


def _row_is_empty(row: Dict[str, Any]) -> bool:
    for v in row.values():
        if v is not None and str(v).strip() != "":
            return False
    return True


def _required_column_keys(data_type: str) -> List[str]:
    """表头校验：部分模板含大量可选列，上传时只要求最小必填列存在。"""
    if data_type == "routings":
        return [
            "code",
            "name",
            "product_code",
            "version",
            "is_active",
            "location",
            "description",
        ]
    if data_type == "resources":
        return ["code", "name"]
    return [spec[0] for spec in FIELD_SPECS[data_type]]


def _parse_upload_sheet(content: bytes, data_type: str) -> List[Tuple[int, Dict[str, Any]]]:
    if data_type not in DATA_TYPES:
        raise ValueError("不支持的数据类型")
    bio = io.BytesIO(content)
    wb = load_workbook(bio, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0
    if max_row < 1 or max_col < 1:
        return []
    headers = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    alias = _alias_to_internal(data_type)
    col_map: Dict[str, int] = {}
    for idx, h in enumerate(headers, start=1):
        key = alias.get(_norm_header(h))
        if key:
            col_map[key] = idx
    required = _required_column_keys(data_type)
    missing = [k for k in required if k not in col_map]
    if missing:
        raise ValueError(f"缺少列: {', '.join(missing)}")
    out: List[Tuple[int, Dict[str, Any]]] = []
    for row_idx in range(2, max_row + 1):
        rowd = _map_row(ws, row_idx, col_map)
        if _row_is_empty(rowd):
            continue
        out.append((row_idx, rowd))
    return out


IMPORTERS = {
    "locations": _import_locations,
    "work_centers": _import_work_centers,
    "resources": _import_resources,
    "shifts": _import_shifts,
    "products": _import_products,
    "routings": _import_routings,
    "routing_operations": _import_routing_operations,
    "setup_groups": _import_setup_groups,
    "product_setup_groups": _import_product_setup_groups,
    "setup_matrix": _import_setup_matrix,
    "production_orders": _import_production_orders,
}


@router.get("/templates")
def list_templates(
    locale: Optional[str] = Query(None),
    accept_language: Optional[str] = Header(None, alias="Accept-Language"),
    user: models.User = Depends(require_auth),
):
    loc = _norm_locale(locale, accept_language)
    use_en = loc == "en-US"
    items = []
    labels = {
        "locations": ("位置", "Locations"),
        "work_centers": ("工作中心", "Work centers"),
        "resources": ("资源", "Resources"),
        "shifts": ("班次", "Shifts"),
        "products": ("产品", "Products"),
        "routings": ("工艺路线", "Routings"),
        "routing_operations": ("工艺路线工序", "Routing operations"),
        "setup_groups": ("切换组", "Setup groups"),
        "product_setup_groups": ("产品切换组", "Product setup groups"),
        "setup_matrix": ("切换矩阵", "Setup matrix"),
        "production_orders": ("生产订单", "Production orders"),
    }
    for dt in DATA_TYPES:
        z, e = labels[dt]
        items.append({"id": dt, "name": e if use_en else z})
    return {"locale": loc, "types": items}


@router.get("/templates/{data_type}")
def download_template(
    data_type: str,
    locale: Optional[str] = Query(None),
    accept_language: Optional[str] = Header(None, alias="Accept-Language"),
    user: models.User = Depends(require_auth),
):
    if data_type not in DATA_TYPES:
        raise HTTPException(status_code=404, detail="不支持的数据类型")
    loc = _norm_locale(locale, accept_language)
    data = _build_workbook(data_type, loc)
    filename = f"{data_type}_template.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    data_type: str = Form(...),
    db: Session = Depends(get_db),
    user: models.User = Depends(require_auth),
):
    if data_type not in DATA_TYPES:
        raise HTTPException(status_code=400, detail="不支持的数据类型")
    name = (file.filename or "").lower()
    if not name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")
    raw = await file.read()
    try:
        rows = _parse_upload_sheet(raw, data_type)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    importer = IMPORTERS[data_type]
    try:
        ok, fail, errors = importer(db, rows)
        db.commit()
    except Exception:
        db.rollback()
        raise
    return {"success": fail == 0, "imported": ok, "failed": fail, "errors": errors[:100]}


@router.post("/clear")
def clear_data(
    body: ClearBody,
    db: Session = Depends(get_db),
    user: models.User = Depends(require_admin),
):
    if body.all:
        to_clear = [k for k, _ in CLEAR_STEPS]
    elif body.data_types:
        req = set(body.data_types)
        bad = req - set(DATA_TYPES)
        if bad:
            raise HTTPException(status_code=400, detail=f"未知类型: {', '.join(sorted(bad))}")
        to_clear = _expand_clear_types(req)
    else:
        raise HTTPException(status_code=400, detail="请指定 data_types 或 all: true")

    counts_before: Dict[str, int] = {}
    try:
        for key, model in CLEAR_STEPS:
            if key not in to_clear:
                continue
            n = db.query(model).count()
            counts_before[key] = n
            db.query(model).delete(synchronize_session=False)
        db.commit()
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="清空失败，已回滚")
    return {"cleared": to_clear, "deleted_counts": counts_before}
